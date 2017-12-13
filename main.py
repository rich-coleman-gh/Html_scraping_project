import urllib.request  as urllib2 #import the library used to query a website
from bs4 import BeautifulSoup #import the Beautiful soup functions to parse the data returned from the website
import pandas as pd #import pandas to convert list to data frame
from openpyxl import load_workbook

# INPUT VARIABLES SPECIFIED BY THE USER
input1 = input('Please input the column in excel that you want modified (UPPERCASE ONLY): ')
input2 = input('\nPlease count the number of the table from the web you would like to parse.\nFor example input 9 if you would like to read from the 9th table listed: ')
input3 = input('\nPlease input the number of the column from the table on the web you would like to parse.\nFor example input 3 if you would like to read from the 3rd column: ')
input4 = input('\nPlease input the number of the excel sheet that you would like to modify.\n For example from left to right the sheet tbas would be 1,2,3... accordingly: ')
input5 = input('\nPlease input the name of the file you would like to modify (extension included).\n For example Verisk Model_Send_Excel_2.xlsx: ')
input6 = input('\nPlease input the path where this folder is located on your computer (please include a "/" at the end of the path).\nFor Example ~/Documents/Git/Html_scraping_project/: ')
input7 = input('\nPlease input the url containing the table that you want to parse.\nFor example http://www.verisk.com/press-releases/2017/february/verisk-analytics-inc-reports-fourth-quarter-2016-financial-results.html:  ')

#Convert user input into proper indexes
def excelColumnToIndex(column):
	return ord(column) - 65

def tableFromWebToIndex(index):
	return int(index) - 1

def tableColumnFromWebToIndex(index):
	return int(index) - 1

def excelSheetToIndex(index):
	return int(index) - 1


#Set global variabes to correct values
EXCEL_COLUMN_INDEX = excelColumnToIndex(input1)
TABLE_FROM_WEB_INDEX = tableFromWebToIndex(input2)
TABLE_FROM_WEB_COLUMN_INDEX = tableColumnFromWebToIndex(input3)
EXCEL_SHEET_INDEX = excelSheetToIndex(input4)
FILENAME = input5
PATH = input6
URL = input7


def parseTables(all_tables):
	parsed_tables = []
	for i in range(len(all_tables)-1):
	    table_body = all_tables[i].find('tbody')

	    rows = table_body.find_all('tr')

	    df_temp = []
	    for row in rows:
	        cols =row.find_all('td')
	        cols = [ele.text.strip() for ele in cols]
	        df_temp.append([ele for ele in cols]) #get rid of empty values
	    parsed_tables.append(df_temp)
	return parsed_tables

def loadExcelDoc(sheet_index):
    # Open up Faton Excel file
    xl = pd.ExcelFile(PATH + FILENAME)
    sheets = xl.sheet_names


    ## open up the first sheet and print our data
    df = xl.parse(sheets[sheet_index],header=None)
    #row_index = df.iloc[0][0]

    #df = xl.parse(sheets[sheet_index])
    #df = df.set_index(row_index)

    return df

def main():
	user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'
	headers={'User-Agent':user_agent,}

	request=urllib2.Request(URL,None,headers) #The assembled request
	page = urllib2.urlopen(request)
	data = page.read() # The data


	#Parse the html in the 'page' variable, and store it in Beautiful Soup format
	soup = BeautifulSoup(data,'html.parser')

	#####gather all tables from page into array 
	all_tables = soup.find_all("table", class_="table-blue")

	parsed_tables = parseTables(all_tables)

	df_from_web = pd.DataFrame(parsed_tables[TABLE_FROM_WEB_INDEX])

	df = loadExcelDoc(EXCEL_SHEET_INDEX)

	wb = load_workbook(FILENAME, keep_vba = True)

	wb.get_sheet_names()

	active_sheet = wb.sheetnames[EXCEL_SHEET_INDEX]

	ws = wb[active_sheet]

	# Lets try to match the row index names from the web to the excel doc
	excel_labels = [i.value for i in ws['A']]# we assume that the labels are always in column A
	web_labels = df_from_web.loc[:,0] # we assume that the label is always in the first column of the dataframe

	for i,excel_label in enumerate(excel_labels):
	    for j,web_label in enumerate(web_labels):
	        if excel_label == web_label:
	            #set the cell value in the excel file to match the value found from the web
	            ws[i+1][EXCEL_COLUMN_INDEX].value = df_from_web.loc[j,TABLE_FROM_WEB_COLUMN_INDEX]
	       
	wb.save("temp.xlsx")

#########################################################################
# Lets run our script
main()